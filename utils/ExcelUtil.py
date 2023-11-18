import csv
import logging
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelUtil:
    @staticmethod
    def filter_by_fields(fields: [str] = [], data: [[]] = []):
        """
        1, find all matched column index from data[0] associated with fields
        2, collect matched columns into new data_filtered, then return

        """
        # figure out which columns should be kept
        first_row = data[0]
        column_idx = ExcelUtil.find_column_index_arr_via_fields(first_row, fields)

        logging.info('fields:' + str(fields) + ' @column: ' + str(column_idx))

        # construct new data in [[]]
        data_filtered: [[]] = []
        for row_idx, row in enumerate(data):
            new_row: [] = []
            for col_idx in column_idx:
                new_row.append(row[col_idx])

            data_filtered.append(new_row)

        return data_filtered

    @staticmethod
    def find_column_index_arr_via_fields(first_row: [], fields: [str] = []):
        column_idx: [int] = []
        for f_idx, field in enumerate(fields):
            for c_idx, cell in enumerate(first_row):
                if field == cell:
                    column_idx.append(c_idx)
        return column_idx

    @staticmethod
    def get_data_from_csv(csvFilePath, skipFirst=False, dlr='\t'):
        result = []
        with open(csvFilePath, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=dlr)
            rowNum = 0
            for row in reader:
                rowNum += 1
                if skipFirst and rowNum == 1:
                    continue
                result.append(row)
        logging.info(f'get_data_from_csv file: {csvFilePath}, data: {str(result)}')
        return result

    @staticmethod
    def convert_csv_to_xlsx(csvFilePath, xlsxFilePath, dlr='\t'):
        wb = Workbook()
        ws = wb.active

        with open(csvFilePath, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=dlr)
            for row in reader:
                ws.append(row)

        wb.save(xlsxFilePath)
        logging.info(f'convert_csv_to_xlsx save file to: {xlsxFilePath}')

    @staticmethod
    def write_dict_to_excel(full_file_path: str, data_as_dict: dict[str:[]]):
        if len(data_as_dict) == 0:
            logging.info(f'not data to write: {full_file_path}')
            return

        wb = Workbook()
        # remove the default active sheet.
        wb.remove(wb.active)

        for k, v in data_as_dict.items():
            sheet: Worksheet = wb.create_sheet(k)
            for row in v:
                sheet.append(row)
        # save to target file
        wb.save(full_file_path)

    @staticmethod
    def get_data_from_excel_file(fileName, startAtRow=0, endAtColumn=50, sheet_index=0):
        result = []
        wb = load_workbook(filename=fileName, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        for idx_row, row in enumerate(ws.rows):
            if idx_row >= startAtRow:
                cells = [str(cell.value)
                         if cell.value is not None
                         else '' for idx_cell, cell in enumerate(row) if idx_cell < endAtColumn
                         ]
                if cells and cells[0] is not None:
                    result.append(cells)
                else:
                    break
        logging.info(f'Loaded {len(result)} records from file: {fileName}')
        return result

    @staticmethod
    def create_folder_if_not_existed(filePath):
        os.makedirs(filePath, exist_ok=True)

    @staticmethod
    def get_log_file_path(app) -> str:
        log_folder = os.path.realpath('log')
        ExcelUtil.create_folder_if_not_existed(log_folder)
        return f'{log_folder}/{app}.log'
